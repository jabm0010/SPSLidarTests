from _datetime import datetime
import glob
import subprocess
import os
from openpyxl import Workbook


"""
Returns dictionary of the bounding box coordinates
"""


def loadDirectory(path):
    filesPaths = glob.glob(path + '\\*', recursive=False)

    files = []
    for filePath in filesPaths:
        tuple = ("files", (filePath, open(filePath, 'rb'), 'application/octet-stream'))
        files.append(tuple)

    return files


def writeFile(request, workspaceName, datasetName, nodeId):
    fileName = workspaceName + "_" + datasetName + "_" + nodeId + "_" + str(datetime.now().microsecond) + ".laz"
    path = os.path.join(os.getcwd(),"files", workspaceName, datasetName)
    print(path)
    if not os.path.exists(path):
        os.makedirs(path)
    with open(os.path.join(path,fileName), "wb") as f:
        f.write(request.content)


def mergeFiles():
    inputFiles = "*.laz"
    output = "merge.laz"
    args = ["lasmerge", "-i", inputFiles, "-o", output]
    subprocess.call(args)
    return output


def visualizeFromLasTools():
    fileToVisualize = mergeFiles()
    args = ["lasview", fileToVisualize]
    subprocess.call(args)


def deleteLazFiles():
    files_in_directory = os.listdir(os.getcwd())
    filtered_files = [file for file in files_in_directory if file.endswith(".laz")]
    for file in filtered_files:
        os.remove(file)


def createSheet(database, results):
    workbook = Workbook()

    sheet = workbook.create_sheet(title = database)
    initRow = 1

    for dataset in results.keys():
        sheet.cell(column = 2, row = initRow, value = dataset)
        initRow +=12
        initCol = 4
        for maxSize in results[dataset].keys():
            tmpRow = initRow
            _ = sheet.cell(column=initCol, row=tmpRow-1, value=maxSize)
            for valueToWrite in results[dataset][maxSize]:
                _ = sheet.cell(column=initCol, row=tmpRow, value= valueToWrite)
                tmpRow+=1
            initCol+=1

    workbook.save(database+".xlsx")

